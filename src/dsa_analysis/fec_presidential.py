from __future__ import annotations

import csv
import io
import re
import ssl
import urllib.request
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import certifi
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
FEC_2016_WORKBOOK_URL = (
    "https://www.fec.gov/resources/cms-content/documents/federalelections2016.xlsx"
)
FEC_2016_RESULTS_PAGE = (
    "https://www.fec.gov/introduction-campaign-finance/"
    "election-results-and-voting-information/federal-elections-2016/"
)
DSA_2016_ENDORSEMENT_URL = (
    "https://www.dsausa.org/news/dsa-endorses-bernie-sanders-for-president/"
)
FEC_2020_WORKBOOK_URL = (
    "https://www.fec.gov/resources/cms-content/documents/federalelections2020.xlsx"
)
FEC_2020_RESULTS_PAGE = (
    "https://www.fec.gov/introduction-campaign-finance/"
    "election-results-and-voting-information/federal-elections-2020/"
)
DSA_2020_ENDORSEMENT_URL = (
    "https://www.dsausa.org/bernie-2020-dsa-endorsement-debate-process/"
)


def import_2016_presidential_primaries(
    *,
    endorsements_path: Path = ROOT / "data" / "manual" / "endorsements.csv",
    candidates_path: Path = ROOT / "data" / "manual" / "race_candidates.csv",
    workbook_bytes: bytes | None = None,
) -> tuple[int, int]:
    payload = workbook_bytes if workbook_bytes is not None else _download_workbook()
    endorsements, candidates = parse_2016_presidential_primaries(payload)
    _merge_prefixed_rows(
        endorsements_path,
        endorsements,
        key="endorsement_id",
        prefix="endorsement-bernie-sanders-president-2016-",
    )
    _merge_prefixed_rows(
        candidates_path,
        candidates,
        key="race_candidate_id",
        prefix="us-president-dem-primary-2016-",
    )
    return len(endorsements), len(candidates)


def import_2020_presidential_primaries(
    *,
    endorsements_path: Path = ROOT / "data" / "manual" / "endorsements.csv",
    candidates_path: Path = ROOT / "data" / "manual" / "race_candidates.csv",
    workbook_bytes: bytes | None = None,
) -> tuple[int, int]:
    payload = (
        workbook_bytes
        if workbook_bytes is not None
        else _download_workbook(FEC_2020_WORKBOOK_URL)
    )
    endorsements, candidates = parse_2020_presidential_primaries(payload)
    _merge_prefixed_rows(
        endorsements_path,
        endorsements,
        key="endorsement_id",
        prefix="endorsement-bernie-sanders-president-2020-",
    )
    _merge_prefixed_rows(
        candidates_path,
        candidates,
        key="race_candidate_id",
        prefix="us-president-dem-primary-2020-",
    )
    return len(endorsements), len(candidates)


def parse_2016_presidential_primaries(
    workbook_bytes: bytes,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    return _parse_presidential_primaries(
        workbook_bytes,
        sheet_title="2016 Pres Primary Results",
        year=2016,
        endorsement_date="2015-06-22",
        endorsement_document_id="dsa-national-bernie-sanders-2015",
        endorsement_url=DSA_2016_ENDORSEMENT_URL,
        results_page=FEC_2016_RESULTS_PAGE,
    )


def parse_2020_presidential_primaries(
    workbook_bytes: bytes,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    return _parse_presidential_primaries(
        workbook_bytes,
        sheet_title="10. 2020 Pres Primary Results",
        year=2020,
        endorsement_date="2019-03-21",
        endorsement_document_id="dsa-national-bernie-sanders-2019",
        endorsement_url=DSA_2020_ENDORSEMENT_URL,
        results_page=FEC_2020_RESULTS_PAGE,
    )


def _parse_presidential_primaries(
    workbook_bytes: bytes,
    *,
    sheet_title: str,
    year: int,
    endorsement_date: str,
    endorsement_document_id: str,
    endorsement_url: str,
    results_page: str,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook[sheet_title]
    grouped: dict[tuple[str, str, date], list[dict[str, Any]]] = defaultdict(list)
    for values in worksheet.iter_rows(min_row=2, max_col=13, values_only=True):
        (
            _,
            _fec_id,
            state,
            state_code,
            primary_date,
            first_name,
            last_name,
            display_name,
            _,
            party,
            votes,
            _percentage,
            _footnotes,
        ) = values
        if party != "D" or not state or not state_code or not primary_date or not display_name:
            continue
        election_date = (
            primary_date.date() if isinstance(primary_date, datetime) else primary_date
        )
        if not isinstance(election_date, date):
            continue
        candidate_name = _candidate_name(first_name, last_name, display_name)
        grouped[(str(state), str(state_code), election_date)].append(
            {
                "candidate_name": candidate_name,
                "votes": int(votes or 0),
            }
        )

    endorsements = []
    candidates = []
    for (state, state_code, election_date), rows in sorted(
        grouped.items(), key=lambda item: (item[0][2], item[0][1])
    ):
        collapsed: dict[str, dict[str, Any]] = {}
        for row in rows:
            existing = collapsed.get(row["candidate_name"])
            if existing is None or row["votes"] > existing["votes"]:
                collapsed[row["candidate_name"]] = row
        rows = list(collapsed.values())
        if not any(row["candidate_name"] == "Bernie Sanders" for row in rows):
            continue
        race_id = f"us-president-dem-primary-{year}-{state_code.casefold()}"
        winner_votes = max(row["votes"] for row in rows)
        sanders_votes = next(
            row["votes"] for row in rows if row["candidate_name"] == "Bernie Sanders"
        )
        endorsements.append(
            {
                "endorsement_id": (
                    f"endorsement-bernie-sanders-president-{year}-{state_code.casefold()}"
                ),
                "race_id": race_id,
                "candidate_id": "bernie-sanders",
                "candidate_name": "Bernie Sanders",
                "office": "President",
                "jurisdiction": state,
                "election_date": election_date.isoformat(),
                "primary_party": "Democratic",
                "endorsing_body": "DSA National",
                "endorsement_date": endorsement_date,
                "endorsement_source_document_id": endorsement_document_id,
                "outcome": "Win" if sanders_votes == winner_votes else "Loss",
                "verification_status": "verified",
            }
        )
        for row in rows:
            candidate_name = row["candidate_name"]
            candidates.append(
                {
                    "race_candidate_id": f"{race_id}-{_slug(candidate_name)}",
                    "race_id": race_id,
                    "candidate_id": _slug(candidate_name),
                    "candidate_name": candidate_name,
                    "party": "Democratic",
                    "role": "endorsed"
                    if candidate_name == "Bernie Sanders"
                    else "opponent",
                    "ballot_status": "certified",
                    "outcome": "Win" if row["votes"] == winner_votes else "Loss",
                    "evidence_status": "verified",
                    "source_url": results_page,
                    "notes": (
                        f"Official FEC {year} results compilation supplies the certified "
                        f"{state} Democratic presidential-primary roster and vote totals; "
                        f"DSA National endorsed Sanders on {endorsement_date} "
                        f"({endorsement_url})."
                    ),
                }
            )
    return endorsements, candidates


def _download_workbook(url: str = FEC_2016_WORKBOOK_URL) -> bytes:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "dsa-analysis/0.1 (+source-first academic research)"},
    )
    context = ssl.create_default_context(cafile=certifi.where())
    with urllib.request.urlopen(request, timeout=120, context=context) as response:
        return response.read()


def _candidate_name(first_name: Any, last_name: Any, display_name: Any) -> str:
    if first_name and last_name:
        return " ".join(str(value).strip() for value in (first_name, last_name))
    return str(last_name or display_name).strip()


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")


def _merge_prefixed_rows(
    path: Path,
    generated_rows: Iterable[dict[str, str]],
    *,
    key: str,
    prefix: str,
) -> None:
    generated = list(generated_rows)
    existing = _read_csv(path)
    retained = [row for row in existing if not row.get(key, "").startswith(prefix)]
    rows = retained + generated
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))
